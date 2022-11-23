from typing import List
from openpyxl import *
from TsWriterClass import *
import sys
import os
import datetime

surfix = {
    "KOR"             :  "ko_KR", 
    "US_ENG"          :  "en_US", 
    "US_FRENCH"       :  "fr_CA", 
    "US_SPAIN"        :  "es_US", 
    "US_PORTUGUESE"   :  "pt_US", 
    "CHINA"           :  "zh_CN", 
    "ARABIC"          :  "ar_SA", 
    "UK_ENG"          :  "en_GB", 
    "PORTUGUESE"      :  "pt_PT", 
    "SPAIN"           :  "es_ES", 
    "FRENCH"          :  "fr_FR", 
    "ITALIAN"         :  "it_IT", 
    "GERMAN"          :  "de_DE", 
    "RUSSIAN"         :  "ru_RU", 
    "DUTCH"           :  "nl_NL", 
    "SWEDISH"         :  "sv_SE", 
    "POLISH"          :  "pl_PL", 
    "TURKISH"         :  "tr_TR", 
    "CZECH"           :  "cs_CZ", 
    "DANISH"          :  "da_DK", 
    "SLOVAKIA"        :  "sk_SK", 
    "NORWEGIAN"       :  "no_NO", 
    "HUNGARIAN"       :  "hu_HU", 
    "JP"              :  "jp_JP", 
    "AU ENG"          :  "en_AU", 
}

def writeTS(fileName, locale, category, strCol, dataCol):
    print("Write data to file {name}".format(name=fileName))
    ts = TsWriter(fileName)
    ts.writeHeader()
    ts.addAttributes(["version", "language"], ["2.0", locale])
    ts.openTagWithAttribute("TS")
    ts.openTag("context")
    ts.openInlineTag("name", category)

    for idx in range(1, len(strCol)):
        if not dataCol[idx].value:
            continue
        ts.openTag("message")
        ts.openInlineTag("source", str(strCol[idx].value).strip())
        ts.openInlineTag("translation", str(dataCol[idx].value).strip())
        ts.closeTag()   # close tag message

    ts.closeTag()   # close tag context
    ts.closeTag()   # close tag TS


def main():
    if len(sys.argv) <= 1:
        raise SyntaxError("Need a xlsx file")

    xlsxName = sys.argv[1]
    if not os.path.exists(xlsxName):
        raise FileNotFoundError("{file} not found".format(file=xlsxName))
    if not xlsxName.endswith(".xlsx"):
        raise NameError("It's not .xlsx file")

    list_keys = list(surfix.keys())
    workbook = load_workbook(xlsxName, read_only=False, keep_vba=False, data_only=True, keep_links=False)
    worksheet = workbook[list(workbook.sheetnames)[0]]
    first_row = list(worksheet.rows)[0] 

    for item in first_row:
        if item.value == "APP":
            tmp_col = list(worksheet.columns)[item.column - 1]
            appName = tmp_col[1].value
        elif item.value == "CATEGORY":
            tmp_col = list(worksheet.columns)[item.column - 1]
            category = tmp_col[1].value
        elif item.value == "STR_ID":
            stringIDcol = list(worksheet.columns)[item.column - 1]
        elif not item.value:
            break

    for item in first_row:
        if item.value in list_keys:
            file = "output/" + appName + "/" + appName + "_" + surfix[item.value] + ".ts"
            dataCol = list(worksheet.columns)[item.column - 1]
            writeTS(file, surfix[item.value], category, stringIDcol, dataCol)
        elif not item.value:
            print("Break because of empty")
            break

    workbook.close()

def test():
    workbook = load_workbook("files/home.xlsx", read_only=True)
    worksheet = workbook[list(workbook.sheetnames)[0]]



if __name__ == "__main__":
    start_time = datetime.datetime.now()
    main()
    print(datetime.datetime.now() - start_time)
