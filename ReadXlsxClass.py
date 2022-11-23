from typing import Type
from openpyxl import load_workbook
from io import BytesIO

class ReadXlsx:
    def __init__(self, fileName):
        with open(fileName, "rb") as f:
            in_mem_file = BytesIO(f.read())
        self.__workbook = load_workbook(in_mem_file, read_only=True, keep_vba=False, data_only=False, keep_links=False)
        self.__data = []

    # Private methods

    # Public methods
    def setWorksheetData(self, index: int) -> None:
        if not type(index) is int:
            raise TypeError("Index must be integer")
        ws = self.getSheetAtIndex(index)
        self.__data = [row for row in ws.iter_rows()]
        self.__totalRows = len(self.__data)
        self.__totalColumns = len(self.__data[0])

    def getItemsAtRow(self, row: int) -> list:
        if not type(row) is int:
            raise TypeError("Row must be integer")
        elif not self.__data:
            raise ValueError("Worksheet Data must be set first, please call function: setWorksheetData(index)")
        return list(self.__data[row])

    def getSheetAtIndex(self, index):
        if not type(index) is int:
            raise TypeError("Index must be integer")
        return self.__workbook[list(self.__workbook.sheetnames)[index]]

    def getValueAt(self, row, col):
        if not (type(row) is int and type(col) is int):
            raise TypeError("Row and Column must be integers")
        elif not self.__data:
            raise ValueError("Worksheet Data must be set first, please call function: setWorksheetData(index)")
        return self.getItemsAtRow(row)[col]

    def totalColumns(self) -> int:
        return self.__totalColumns

    def totalRows(self) -> int:
        return self.__totalRows